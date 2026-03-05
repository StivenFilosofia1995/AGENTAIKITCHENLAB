"""
Módulo agente de extracción de facturas.
- Conexión IMAP a Gmail
- Extracción de datos con Claude AI (Anthropic)
- Procesamiento de PDF, DOCX, XLSX adjuntos
- Guardado en Google Sheets
NO contiene rutas FastAPI — eso es responsabilidad de backend.py
"""

from dotenv import load_dotenv
load_dotenv()

import os
import json
import time
import imaplib
import email
from email.header import decode_header
import base64
import io
import re
from datetime import datetime
from typing import Optional

# ── MODELO ────────────────────────────────────────────────────────────────────
# claude-sonnet-4-6: modelo ideal para extracción de facturas DIAN
# → Mayor precisión en campos complejos (rete_iva, rete_ica, XML UBL 2.1)
# → Mejor manejo de PDFs con tablas, imágenes y adjuntos múltiples
# → Mismo costo que Sonnet 4.5, con mejoras en razonamiento estructurado
CLAUDE_MODEL = "claude-sonnet-4-6"

# ── CONSTANTES ────────────────────────────────────────────────────────────────
_BASE_DIR            = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH           = os.getenv("EXCEL_OUTPUT_PATH", os.path.join(_BASE_DIR, "facturas_seguimiento.xlsx"))
PROCESSED_FILE       = os.path.join(_BASE_DIR, "processed_emails.json")
IMAP_HOST            = os.getenv("IMAP_HOST", "imap.gmail.com")
IMAP_PORT            = int(os.getenv("IMAP_PORT", "993"))
IMAP_FOLDER          = os.getenv("IMAP_FOLDER", "INBOX")
EMAIL_USER           = os.getenv("EMAIL_USER", "")
EMAIL_PASS           = os.getenv("EMAIL_PASS", "")
ANTHROPIC_API_KEY    = os.getenv("ANTHROPIC_API_KEY", "")
GOOGLE_SHEETS_ID     = os.getenv("GOOGLE_SHEETS_ID", "")
SERVICE_ACCOUNT_FILE = os.path.join(_BASE_DIR, os.getenv("SERVICE_ACCOUNT_FILE", "service_account.json"))
SHEET_NAME           = "Facturas"

COLUMNS = [
    "Mes", "Fecha Factura", "Número Factura", "Proveedor", "ID", "Número ID",
    "Subtotal", "Descuento", "IVA", "Rete IVA", "Rete ICA", "Impto Consumo",
    "Propina", "Otros Impuestos", "Retención en la fuente", "Administración",
    "Utilidad", "Imprevistos", "Valor Total", "Clasificación", "Estado",
    "Valor Pagado", "Valor Por Pagar", "Fecha Pago", "Cliente",
    "Cotización Inventto", "Observaciones"
]

# ── GOOGLE SHEETS CLIENT ──────────────────────────────────────────────────────
_gspread_client = None
_gspread_spreadsheet = None  # Cache del spreadsheet para evitar open_by_key repetidos
_gspread_worksheets = {}     # Cache de worksheets por nombre {nombre: ws}

MESES = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
         "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]


def _api_call_with_retry(fn, *args, max_retries=5, **kwargs):
    """Ejecuta una llamada a la API de Google con retry exponencial en caso de 429."""
    import gspread
    for attempt in range(max_retries):
        try:
            return fn(*args, **kwargs)
        except gspread.exceptions.APIError as e:
            if "429" in str(e) or "Quota" in str(e):
                wait = (2 ** attempt) * 2  # 2s, 4s, 8s, 16s, 32s
                print(f"⚠️  Google Sheets cuota excedida. Esperando {wait}s antes de reintentar ({attempt+1}/{max_retries})...")
                time.sleep(wait)
            else:
                raise
    raise Exception("❌ Límite de reintentos alcanzado para Google Sheets API")


def _get_spreadsheet():
    """Retorna el spreadsheet cacheado. Llama open_by_key solo una vez por sesión."""
    global _gspread_client, _gspread_spreadsheet
    import gspread
    from google.oauth2.service_account import Credentials

    SCOPES = [
        "https://spreadsheets.google.com/feeds",
        "https://www.googleapis.com/auth/drive",
    ]

    if _gspread_client is None:
        creds_json_env = os.getenv("GOOGLE_CREDENTIALS_JSON")
        if creds_json_env:
            import json as _json
            creds_dict = _json.loads(creds_json_env)
            creds = Credentials.from_service_account_info(creds_dict, scopes=SCOPES)
        else:
            creds = Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
        _gspread_client = gspread.authorize(creds)

    if _gspread_spreadsheet is None:
        _gspread_spreadsheet = _api_call_with_retry(_gspread_client.open_by_key, GOOGLE_SHEETS_ID)

    return _gspread_spreadsheet


def get_sheet(mes_nombre: str = None):
    """Retorna la hoja de Google Sheets del mes especificado, creándola si no existe."""
    global _gspread_worksheets
    import gspread

    if mes_nombre is None:
        mes_nombre = MESES[datetime.now().month - 1]

    # Retornar desde cache si ya se obtuvo antes
    if mes_nombre in _gspread_worksheets:
        return _gspread_worksheets[mes_nombre]

    spreadsheet = _get_spreadsheet()

    try:
        ws = _api_call_with_retry(spreadsheet.worksheet, mes_nombre)
    except gspread.WorksheetNotFound:
        ws = _api_call_with_retry(spreadsheet.add_worksheet,
                                  title=mes_nombre, rows=1000, cols=len(COLUMNS))
        _api_call_with_retry(ws.append_row, COLUMNS)
        print(f"✅ Hoja '{mes_nombre}' creada con encabezados")

    _gspread_worksheets[mes_nombre] = ws
    return ws


def get_all_monthly_worksheets():
    """Obtiene todas las hojas mensuales existentes en UNA sola llamada a la API."""
    global _gspread_worksheets
    spreadsheet = _get_spreadsheet()
    # Una sola llamada a la API para listar todas las hojas
    all_ws = _api_call_with_retry(spreadsheet.worksheets)
    result = []
    for ws in all_ws:
        if ws.title in MESES:
            _gspread_worksheets[ws.title] = ws  # Actualizar cache
            result.append(ws)
    return result


def setup_sheets():
    """Verifica conexión a Google Sheets y crea hoja del mes actual si no existe."""
    try:
        mes_actual = MESES[datetime.now().month - 1]
        ws = get_sheet(mes_actual)  # Usa cache interno — NO repite open_by_key
        print(f"✅ Google Sheets conectado: Hoja '{mes_actual}'")
    except Exception as e:
        print(f"❌ Error conectando a Google Sheets: {e}")
        raise


def _load_processed_ids() -> set:
    try:
        if os.path.exists(PROCESSED_FILE):
            with open(PROCESSED_FILE) as f:
                return set(json.load(f))
    except Exception:
        pass
    return set()


def _save_processed_ids(ids: set):
    try:
        with open(PROCESSED_FILE, "w") as f:
            json.dump(list(ids), f)
    except Exception as e:
        print(f"⚠️ No se pudo guardar IDs procesados: {e}")


# ── IMAP ──────────────────────────────────────────────────────────────────────
def connect_imap() -> imaplib.IMAP4_SSL:
    mail = imaplib.IMAP4_SSL(IMAP_HOST, IMAP_PORT)
    mail.login(EMAIL_USER, EMAIL_PASS)
    mail.select(IMAP_FOLDER)
    print(f"✅ Conectado a {IMAP_HOST} como {EMAIL_USER}")
    return mail


def _parse_imap_folder_name(raw_line: str) -> str:
    """Extrae el nombre de carpeta de una línea LIST de IMAP.

    Formatos posibles que devuelve Gmail:
      (\\HasNoChildren) "/" "[Gmail]/Todos"
      (\\HasNoChildren \\All) "/" "[Gmail]/All Mail"
      (\\HasNoChildren) "/" INBOX
    Retorna el nombre SIN comillas externas, listo para _imap_select().
    """
    # Separar por el delimitador "/" o NIL
    import re as _re
    # Buscar el último token: puede estar entre comillas o sin ellas
    m = _re.search(r'"/" "(.+)"$', raw_line)
    if m:
        return m.group(1)
    m = _re.search(r'"/" ([^\s"]+)$', raw_line)
    if m:
        return m.group(1)
    # NIL como separador
    m = _re.search(r'NIL "(.+)"$', raw_line)
    if m:
        return m.group(1)
    m = _re.search(r'NIL ([^\s"]+)$', raw_line)
    if m:
        return m.group(1)
    return ""


def _imap_select(mail: imaplib.IMAP4_SSL, folder_name: str) -> bool:
    """Selecciona una carpeta IMAP con el quoting correcto.

    imaplib NO hace quoting automático. Los nombres con espacios o corchetes
    deben enviarse entre comillas dobles en el protocolo IMAP.
    Esta función siempre usa comillas para máxima compatibilidad.
    """
    # Limpiar comillas previas para no doblarlas
    clean = folder_name.strip().strip('"')
    try:
        status, _ = mail.select(f'"{clean}"', readonly=True)
        return status == 'OK'
    except Exception:
        return False


def _find_all_mail_folder(mail: imaplib.IMAP4_SSL) -> str:
    """Descubre el nombre real de la carpeta AllMail/Todos usando LIST.

    Retorna el nombre limpio (sin comillas) o None si no se encuentra.
    El caller debe usar _imap_select() para seleccionarla.

    Prioridades:
    1. Carpeta con atributo IMAP \\All  (RFC 6154 — la más fiable)
    2. Carpeta cuyo nombre contenga palabras clave multiidioma
    3. None → el caller usará INBOX como fallback
    """
    try:
        status, folder_list = mail.list('""', '*')
        if status != 'OK' or not folder_list:
            print("⚠️  No se pudo listar carpetas IMAP")
            return None

        keywords_allmail = ['all mail', 'todos', 'tout', 'alle', 'tutti', 'tous']
        candidate_by_keyword = None

        for raw in folder_list:
            if not raw:
                continue
            decoded = raw.decode('utf-8', errors='replace') if isinstance(raw, bytes) else str(raw)

            name = _parse_imap_folder_name(decoded)
            if not name:
                continue

            # Prioridad 1: atributo \All
            if r'\All' in decoded or r'\all' in decoded:
                print(f"📁 All Mail encontrado por atributo \\All: {name}")
                return name

            # Prioridad 2: palabra clave en el nombre
            if candidate_by_keyword is None:
                name_lower = name.lower()
                for kw in keywords_allmail:
                    if kw in name_lower:
                        candidate_by_keyword = name
                        break

        if candidate_by_keyword:
            print(f"📁 All Mail encontrado por nombre: {candidate_by_keyword}")
            return candidate_by_keyword

    except Exception as e:
        print(f"⚠️  Error listando carpetas IMAP: {e}")

    return None


def search_invoice_emails(mail: imaplib.IMAP4_SSL) -> list:
    """Devuelve los IDs de los 500 correos más recientes desde UN solo folder."""
    all_mail = _find_all_mail_folder(mail)
    folders_to_try = [all_mail] if all_mail else []
    folders_to_try.append('INBOX')

    for folder in folders_to_try:
        if not folder:
            continue
        try:
            ok = _imap_select(mail, folder)
            if not ok:
                continue
            _, data = mail.search(None, "ALL")
            if not data or not data[0]:
                continue
            ids = [uid.decode() for uid in data[0].split() if uid]
            if ids:
                recientes = sorted(ids, key=lambda x: int(x) if x.isdigit() else 0)[-500:]
                print(f"📬 {len(recientes)} correos para analizar [{folder}]")
                return recientes
        except Exception:
            continue
    return []


# Mapa de meses español → número
_MES_NUM = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4,
    "mayo": 5, "junio": 6, "julio": 7, "agosto": 8,
    "septiembre": 9, "octubre": 10, "noviembre": 11, "diciembre": 12
}
# Nombres en inglés para IMAP (formato requerido: "01-Jan-2026")
_MES_EN = {
    1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr", 5: "May", 6: "Jun",
    7: "Jul", 8: "Aug", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec"
}


def search_emails_by_month(mail: imaplib.IMAP4_SSL, mes_nombre: str, year: int = None) -> list:
    """Busca TODOS los correos del mes usando IMAP SINCE/BEFORE.

    Descubre el folder All Mail dinámicamente para soportar Gmail en cualquier
    idioma (español → Todos, inglés → All Mail, etc.).
    """
    mes_nombre = mes_nombre.lower().strip()
    mes_num = _MES_NUM.get(mes_nombre)
    if not mes_num:
        print(f"❌ Mes no reconocido: {mes_nombre}")
        return []

    if year is None:
        year = datetime.now().year

    mes_siguiente  = mes_num + 1 if mes_num < 12 else 1
    year_siguiente = year if mes_num < 12 else year + 1
    fecha_inicio   = f"01-{_MES_EN[mes_num]}-{year}"
    fecha_fin      = f"01-{_MES_EN[mes_siguiente]}-{year_siguiente}"
    criterio       = f'(SINCE "{fecha_inicio}" BEFORE "{fecha_fin}")'

    all_mail_folder = _find_all_mail_folder(mail)
    folders_to_try = [all_mail_folder] if all_mail_folder else []
    folders_to_try.append('INBOX')

    for folder in folders_to_try:
        if not folder:
            continue
        try:
            ok = _imap_select(mail, folder)
            if not ok:
                continue
            _, data = mail.search(None, criterio)
            if not data or not data[0]:
                print(f"📭 0 correos en {mes_nombre.capitalize()} {year} [{folder}]")
                continue
            ids = [uid.decode() for uid in data[0].split() if uid]
            if ids:
                print(f"📬 {len(ids)} correos en {mes_nombre.capitalize()} {year} [{folder}]")
                return sorted(ids, key=lambda x: int(x) if x.isdigit() else 0)
        except Exception as ex:
            print(f"⚠️  Folder {folder} error: {ex}")
            continue

    print(f"📭 No se encontraron correos en {mes_nombre.capitalize()} {year}")
    return []


def process_emails_for_month(mes_nombre: str, year: int = None):
    """Procesa TODOS los correos de un mes específico."""
    if not EMAIL_USER or not EMAIL_PASS:
        print("❌ Credenciales de correo no configuradas en .env")
        return
    if not ANTHROPIC_API_KEY:
        print("❌ ANTHROPIC_API_KEY no configurada en .env")
        return

    mes_nombre_cap = mes_nombre.strip().capitalize()
    if year is None:
        year = datetime.now().year

    print(f"🗓️  Iniciando procesamiento del mes: {mes_nombre_cap} {year}")
    print(f"🤖 Modelo de extracción: {CLAUDE_MODEL}")
    setup_sheets()

    try:
        mail = connect_imap()
    except Exception as e:
        print(f"❌ Error de conexión IMAP: {e}")
        return

    email_ids = search_emails_by_month(mail, mes_nombre, year)
    if not email_ids:
        print(f"📭 No se encontraron correos en {mes_nombre_cap} {year}")
        try:
            mail.logout()
        except Exception:
            pass
        return

    total = len(email_ids)
    facturas_encontradas = 0
    omitidos = 0
    errores = 0
    seen_msgids: set = set()

    for idx, eid in enumerate(email_ids, 1):
        try:
            try:
                raw_result = mail.fetch(eid, "(RFC822)")
                status_f, raw_data = raw_result
            except (imaplib.IMAP4.abort, imaplib.IMAP4.error, OSError):
                print(f"⚠️  Sesión IMAP expirada, reconectando...")
                try:
                    mail.logout()
                except Exception:
                    pass
                mail = connect_imap()
                search_emails_by_month(mail, mes_nombre, year)
                status_f, raw_data = mail.fetch(eid, "(RFC822)")

            if status_f != 'OK' or not raw_data or raw_data[0] is None:
                print(f"⏭️  No se pudo obtener correo {eid}")
                omitidos += 1
                continue

            raw_bytes = None
            for part in raw_data:
                if isinstance(part, tuple) and len(part) >= 2 and isinstance(part[1], bytes):
                    raw_bytes = part[1]
                    break
            if not raw_bytes:
                omitidos += 1
                continue

            msg    = email.message_from_bytes(raw_bytes)
            asunto = _decode_str(msg.get("Subject", "(sin asunto)"))
            from_  = _decode_str(msg.get("From", ""))

            msg_id = str(msg.get("Message-ID", "")).strip()
            if msg_id and msg_id in seen_msgids:
                omitidos += 1
                continue
            if msg_id:
                seen_msgids.add(msg_id)

            body, attachments = _get_email_content(msg)
            if not body.strip() and not attachments:
                omitidos += 1
                continue

            print(f"[{idx}/{total}] 🔍 {asunto[:60]}")

            # ── PRE-FILTRO LOCAL: evita gastar tokens en no-facturas ─────────
            body_preview = body[:300] if body else ""
            if _should_skip_email(asunto, body_preview):
                print(f"[{idx}/{total}] ⏭️  Omitido por pre-filtro: {asunto[:60]}")
                omitidos += 1
                continue
            # ─────────────────────────────────────────────────────────────────

            datos = _extract_with_ai(asunto, body, attachments)
            if datos:
                if not datos.get("mes") or str(datos.get("mes")).upper() in ("", "N/A"):
                    datos["mes"] = mes_nombre_cap
                guardada = save_to_sheets(datos, from_, forced_mes=mes_nombre_cap)
                if guardada:
                    facturas_encontradas += 1
        except Exception as e:
            import traceback
            print(f"⚠️  Error procesando correo {eid}: {e}")
            print(traceback.format_exc())
            errores += 1
            continue

    try:
        mail.logout()
    except Exception:
        pass

    print(
        f"🏁 {mes_nombre_cap} {year} completado: "
        f"{facturas_encontradas} facturas guardadas | "
        f"{total} correos analizados | "
        f"{errores} errores | {omitidos} omitidos"
    )


# ── EXTRACCIÓN DE TEXTO ───────────────────────────────────────────────────────
def _decode_str(s) -> str:
    if not s:
        return ""
    parts = decode_header(s)
    result = []
    for part, enc in parts:
        if isinstance(part, bytes):
            result.append(part.decode(enc or "utf-8", errors="replace"))
        else:
            result.append(str(part))
    return " ".join(result)


def _extract_pdf_text(data: bytes) -> str:
    """Extrae texto de PDF. Usa pdfplumber (tablas) + PyPDF2 fallback."""
    text = ""
    try:
        import pdfplumber
        with pdfplumber.open(io.BytesIO(data)) as pdf:
            parts = []
            for page in pdf.pages:
                pg_text = page.extract_text() or ""
                tables = page.extract_tables() or []
                for table in tables:
                    for row in table:
                        if row:
                            line = " | ".join(str(c).strip() for c in row if c is not None and str(c).strip())
                            if line:
                                parts.append(line)
                if pg_text.strip():
                    parts.append(pg_text)
            text = "\n".join(parts).strip()
    except Exception:
        pass
    if not text:
        try:
            import PyPDF2
            reader = PyPDF2.PdfReader(io.BytesIO(data))
            text = "\n".join(p.extract_text() or "" for p in reader.pages)
        except Exception as e:
            text = f"[Error PDF: {e}]"
    return text


def _extract_docx_text(data: bytes) -> str:
    try:
        from docx import Document
        doc = Document(io.BytesIO(data))
        return "\n".join(p.text for p in doc.paragraphs)
    except Exception as e:
        return f"[Error DOCX: {e}]"


def _extract_xlsx_text(data: bytes) -> str:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        lines = []
        for ws in wb.worksheets:
            lines.append(f"=== Hoja: {ws.title} ===")
            for row in ws.iter_rows(values_only=True):
                row_vals = [str(c).strip() for c in row if c is not None and str(c).strip()]
                if row_vals:
                    lines.append(" | ".join(row_vals))
        return "\n".join(lines[:300])
    except Exception as e:
        return f"[Error XLSX: {e}]"


def _extract_xml_text(data: bytes) -> str:
    """Extrae campos clave de XML DIAN UBL 2.1 + fallback genérico."""
    try:
        import xml.etree.ElementTree as ET
        raw_str = data.decode("utf-8", errors="replace")
        root = ET.fromstring(raw_str)

        def _t(elem, *tags):
            for tag in tags:
                for ns_prefix in ("", "{urn:oasis:names:specification:ubl:schema:xsd:Invoice-2}",
                                  "{urn:oasis:names:specification:ubl:schema:xsd:CommonBasicComponents-2}",
                                  "{urn:oasis:names:specification:ubl:schema:xsd:CommonAggregateComponents-2}"):
                    e = root.find(f".//{ns_prefix}{tag}")
                    if e is not None and e.text and e.text.strip():
                        return e.text.strip()
            return None

        DIAN_FIELDS = [
            ("Número Factura",       ["ID", "InvoiceID", "Folio"]),
            ("Fecha Factura",        ["IssueDate", "FechaEmision"]),
            ("Proveedor NIT",        ["CompanyID", "TaxSchemeID"]),
            ("Proveedor Nombre",     ["RegistrationName", "Name", "PartyName"]),
            ("Subtotal",             ["LineExtensionAmount", "SubTotal"]),
            ("IVA",                  ["TaxAmount"]),
            ("Valor Total",          ["TaxInclusiveAmount", "PayableAmount", "Total"]),
            ("Moneda",               ["DocumentCurrencyCode", "Moneda"]),
            ("Tipo Factura",         ["InvoiceTypeCode"]),
        ]

        targeted = []
        for label, tags in DIAN_FIELDS:
            val = _t(root, *tags)
            if val:
                targeted.append(f"{label}: {val}")

        generic = []
        for elem in root.iter():
            tag = elem.tag.split("}")[-1] if "}" in elem.tag else elem.tag
            if elem.text and elem.text.strip():
                generic.append(f"{tag}: {elem.text.strip()}")
            for ak, av in elem.attrib.items():
                ak = ak.split("}")[-1] if "}" in ak else ak
                generic.append(f"{tag}[{ak}]: {av}")

        combined = "\n".join(targeted) + "\n\n--- XML completo ---\n" + "\n".join(generic[:300])
        return combined
    except Exception as e:
        return f"[Error XML: {e}]"


def _extract_zip_text(data: bytes) -> str:
    """Extrae texto de archivos dentro de un ZIP."""
    try:
        import zipfile
        results = []
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            for name in zf.namelist():
                lower = name.lower()
                try:
                    file_data = zf.read(name)
                except Exception:
                    continue
                if lower.endswith(".pdf"):
                    results.append(f"[ZIP/{name}]\n" + _extract_pdf_text(file_data))
                elif lower.endswith(".xlsx") or lower.endswith(".xls"):
                    results.append(f"[ZIP/{name}]\n" + _extract_xlsx_text(file_data))
                elif lower.endswith(".docx"):
                    results.append(f"[ZIP/{name}]\n" + _extract_docx_text(file_data))
                elif lower.endswith(".xml"):
                    results.append(f"[ZIP/{name}]\n" + _extract_xml_text(file_data))
                elif lower.endswith(".txt"):
                    results.append(f"[ZIP/{name}]\n" + file_data.decode("utf-8", errors="replace")[:2000])
        return "\n\n".join(results) if results else "[ZIP vacío o sin archivos reconocibles]"
    except Exception as e:
        return f"[Error ZIP: {e}]"


def _extract_links_from_body(body: str) -> list:
    """Extrae URLs de un cuerpo de correo e intenta descargar documentos enlazados."""
    import urllib.request
    url_pattern = re.compile(r'https?://[^\s<>"]+', re.IGNORECASE)
    urls = url_pattern.findall(body)
    downloaded = []
    doc_exts = (".pdf", ".xml", ".xlsx", ".xls", ".zip", ".docx")
    seen = set()
    for url in urls:
        url_clean = url.rstrip(".,);'\"")
        if any(url_clean.lower().endswith(ext) for ext in doc_exts) and url_clean not in seen:
            seen.add(url_clean)
            try:
                req = urllib.request.Request(url_clean, headers={"User-Agent": "Mozilla/5.0"})
                with urllib.request.urlopen(req, timeout=8) as resp:
                    raw = resp.read()
                ext = url_clean.split("?")[0].lower()
                if ext.endswith(".pdf"):
                    downloaded.append(f"[URL: {url_clean}]\n" + _extract_pdf_text(raw))
                elif ext.endswith(".xml"):
                    downloaded.append(f"[URL: {url_clean}]\n" + _extract_xml_text(raw))
                elif ext.endswith(".xlsx") or ext.endswith(".xls"):
                    downloaded.append(f"[URL: {url_clean}]\n" + _extract_xlsx_text(raw))
                elif ext.endswith(".zip"):
                    downloaded.append(f"[URL: {url_clean}]\n" + _extract_zip_text(raw))
                elif ext.endswith(".docx"):
                    downloaded.append(f"[URL: {url_clean}]\n" + _extract_docx_text(raw))
                print(f"🔗 Descargado adjunto desde URL: {url_clean}")
            except Exception as e:
                print(f"⚠️  No se pudo descargar {url_clean}: {e}")
    return downloaded


def _get_email_content(msg) -> tuple[str, list]:
    """Retorna (cuerpo_texto, lista_de_textos_adjuntos)."""
    body = ""
    attachments = []

    for part in msg.walk():
        ct = part.get_content_type()
        cd = str(part.get("Content-Disposition", ""))

        if "attachment" in cd or ct in (
            "application/pdf",
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            "application/zip", "application/x-zip-compressed",
            "application/xml", "text/xml",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/pkcs7-mime",
            "image/png", "image/jpeg", "image/jpg", "image/gif", "image/webp",
        ):
            filename = part.get_filename() or ""
            payload = part.get_payload(decode=True)
            if not payload:
                continue
            fname_lower = filename.lower()
            if fname_lower.endswith(".pdf") or ct == "application/pdf":
                attachments.append(f"[PDF: {filename}]\n" + _extract_pdf_text(payload))
                print(f"📎 PDF adjunto procesado: {filename}")
            elif fname_lower.endswith(".docx"):
                attachments.append(f"[DOCX: {filename}]\n" + _extract_docx_text(payload))
                print(f"📎 DOCX adjunto procesado: {filename}")
            elif fname_lower.endswith(".xlsx") or fname_lower.endswith(".xls"):
                attachments.append(f"[XLSX: {filename}]\n" + _extract_xlsx_text(payload))
                print(f"📎 XLSX adjunto procesado: {filename}")
            elif fname_lower.endswith(".xml") or fname_lower.endswith(".p7m") or ct in ("application/xml", "text/xml", "application/pkcs7-mime"):
                xml_bytes = payload
                if fname_lower.endswith(".p7m"):
                    idx = payload.find(b"<?xml")
                    if idx != -1:
                        xml_bytes = payload[idx:]
                attachments.append(f"[XML: {filename}]\n" + _extract_xml_text(xml_bytes))
                print(f"📎 XML adjunto procesado: {filename}")
            elif fname_lower.endswith(".zip") or ct in ("application/zip", "application/x-zip-compressed"):
                attachments.append(f"[ZIP: {filename}]\n" + _extract_zip_text(payload))
                print(f"📎 ZIP adjunto procesado: {filename}")
            elif ct.startswith("image/") or fname_lower.endswith((".png", ".jpg", ".jpeg", ".gif", ".webp")):
                img_b64 = base64.b64encode(payload).decode()
                media_type = ct if ct.startswith("image/") else "image/png"
                attachments.append(("__IMAGE__", media_type, img_b64, filename))
                print(f"📎 Imagen adjunta registrada: {filename}")

        elif ct == "text/plain" and "attachment" not in cd:
            try:
                body += part.get_payload(decode=True).decode(
                    part.get_content_charset() or "utf-8", errors="replace"
                )
            except Exception:
                pass
        elif ct == "text/html" and "attachment" not in cd:
            try:
                from bs4 import BeautifulSoup
                raw = part.get_payload(decode=True).decode(
                    part.get_content_charset() or "utf-8", errors="replace"
                )
                html_text = BeautifulSoup(raw, "html.parser").get_text(separator="\n")
                if html_text.strip() and html_text.strip() not in body:
                    body += "\n" + html_text
            except Exception:
                pass

    if body:
        linked_docs = _extract_links_from_body(body)
        attachments.extend(linked_docs)

    return body, attachments


# ── INTELIGENCIA ARTIFICIAL (Claude Sonnet 4.6) ───────────────────────────────
JSON_SCHEMA = """
{
  "numero_factura": "(string — número/código de la factura)",
  "proveedor": "(string — razón social del emisor/vendedor)",
  "fecha_factura": "DD/MM/YYYY",
  "id_tipo": "NIT | CC | CE | RUT | PASAPORTE",
  "numero_id": "(NIT o cédula del proveedor, sin dígito verificador separado)",
  "subtotal": 0.0,
  "descuento": 0.0,
  "iva": 0.0,
  "rete_iva": 0.0,
  "rete_ica": 0.0,
  "impto_consumo": 0.0,
  "propina": 0.0,
  "otros_impuestos": 0.0,
  "retencion_fuente": 0.0,
  "administracion": 0.0,
  "utilidad": 0.0,
  "imprevistos": 0.0,
  "valor_total": 0.0,
  "clasificacion": "Servicios | Productos | Mixto",
  "estado": "PENDIENTE | PAGADA | VENCIDA",
  "valor_pagado": 0.0,
  "valor_por_pagar": 0.0,
  "fecha_pago": "DD/MM/YYYY o N/A",
  "cliente": "(string — razón social del receptor/comprador)",
  "cotizacion": "(número de cotización si aparece, si no N/A)",
  "observaciones": "(notas relevantes, descripción del servicio/producto)"
}"""


def _smart_truncate(text: str, max_chars: int) -> str:
    """Toma el inicio y el final del texto para no perder datos clave al final del doc."""
    if len(text) <= max_chars:
        return text
    half = max_chars // 2
    return text[:half] + "\n...[CONTENIDO OMITIDO]...\n" + text[-half:]



# ── PRE-FILTRO LOCAL (sin tokens de IA) ──────────────────────────────────────
_SKIP_SUBJECTS = [
    # Certificados y solicitudes administrativas
    "certificado de retenci",
    "certificado tributar",
    "certificado de ingreso",
    "solicitud de certificado",
    "solicitud de informaci",
    "solicitud de cotizaci",
    "declaraci",
    "informe de gestion",
    # Cuentas de cobro y honorarios (excluidas explícitamente)
    "cuenta de cobro",
    "cobro de honorarios",
    "honorarios profesionales",
    # Links / notificaciones de pago (NO la factura en sí)
    "link de pago",
    "enlace de pago",
    "boton de pago",
    "botón de pago",
    "notificaci",               # notificación PSE / Nequi / Bancolombia
    "pago exitoso",
    "pago recibido",
    "transacci",
    "extracto",
    # Seguridad / auth
    "cambio de clave",
    "cambio de contrase",
    "restablecimiento",
    "recupera tu",
    "verificaci",
    "confirma tu correo",
    "activa tu cuenta",
    "inicio de sesi",
    "two-factor",
    "autenticaci",
    # Marketing y boletines
    "boletin",
    "boletín",
    "newsletter",
    "oferta especial",
    "promocion",
    "promoción",
    "unsubscribe",
    "darse de baja",
    # Soporte y encuestas
    "ticket #",
    "caso #",
    "soporte técnico",
    "soporte tecnico",
    "mesa de ayuda",
    "encuesta",
    "califica tu experiencia",
    "satisfacci",
]

# Palabras que SÍ indican factura electrónica DIAN (whitelist — anulan el skip)
_INVOICE_KEYWORDS = [
    "factura electronica",
    "factura electrónica",
    "factura de venta",
    "fact-", "fact.", "fe-", "fv-", "fes-",
    "cufe", "cude",
    "nota crédito", "nota credito",
    "nota débito",  "nota debito",
    "invoice",
]

def _should_skip_email(subject: str, body_preview: str = "") -> bool:
    """Pre-filtro LOCAL: descarta correos que definitivamente no son facturas
    sin gastar ningún token de IA.

    Reglas (en orden de prioridad):
    1. Si el asunto contiene una palabra de factura DIAN → procesar siempre.
    2. Si el asunto contiene una palabra de la blacklist → omitir.
    3. Sin indicadores claros → dejar pasar (Claude decide).
    """
    subj_low = subject.lower().strip()

    # 1. Whitelist: factura DIAN confirmada → nunca saltar
    for kw in _INVOICE_KEYWORDS:
        if kw in subj_low:
            return False

    # 2. Blacklist: documentos que el usuario quiere excluir
    for kw in _SKIP_SUBJECTS:
        if kw in subj_low:
            return True

    return False


def _extract_with_ai(subject: str, body: str, attachments: list) -> Optional[dict]:
    """Usa claude-sonnet-4-6 para extraer datos de la factura. Retorna dict o None.

    Rate limiting: espera 3 s entre llamadas (≈20 req/min) y reintenta 3 veces
    con backoff exponencial si llega un 429 de Anthropic.
    """
    import anthropic, time as _time

    # Throttle base: garantiza ≈ 3 s entre llamadas a la API
    _time.sleep(3)

    for attempt in range(3):   # hasta 3 intentos por correo
      try:
        client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)

        text_parts  = [f"ASUNTO DEL CORREO: {subject}"]
        image_parts = []

        # Reducir cuerpo de 12000 a 6000 chars: la mayoría de XMLs DIAN
        # tienen los datos clave en los primeros 5-6 kB.
        body_clean = _smart_truncate(body.strip(), 6000)
        if body_clean:
            text_parts.append(f"CUERPO DEL CORREO:\n{body_clean}")

        for i, att in enumerate(attachments, 1):
            if isinstance(att, tuple) and att[0] == "__IMAGE__":
                _, media_type, b64, fname = att
                image_parts.append((media_type, b64, fname))
            else:
                # Adjuntos de texto: reducir a 6000 chars también
                att_text = _smart_truncate(str(att), 6000)
                text_parts.append(f"ADJUNTO {i}:\n{att_text}")

        full_text = "\n\n".join(text_parts)

        prompt = f"""Eres un experto en contabilidad colombiana y facturas electrónicas DIAN.
Analiza el contenido que se te da (cuerpo del correo, adjuntos PDF, XML DIAN, XLSX, imágenes).

═══════════════════════════════════════════════════════
RESPONDE NO_ES_FACTURA — sin excepción — si el documento es:
═══════════════════════════════════════════════════════
- Solicitud de certificado (retención, tributario, ingresos, etc.)
- Solicitud de información o formulario
- Cuenta de cobro o cobro de honorarios
- Link / botón de pago (PSE, Wompi, PayU, Nequi, Bancolombia, etc.)
- Cambio de contraseña / clave / verificación 2FA
- Cotización sin número de factura DIAN oficial
- Extracto bancario
- Publicidad, boletín o correo de soporte sin factura adjunta

═══════════════════════════════════════════════════════
PROCESA ÚNICAMENTE:
═══════════════════════════════════════════════════════
FACTURA ELECTRÓNICA DIAN (XML UBL 2.1, PDF, imagen, XLSX)
con CUFE o número de factura oficial (FE-, FV-, FES-, FACT-, etc.).

═══════════════════════════════════════════════════════
REGLAS DE EXTRACCIÓN (solo si es factura DIAN válida)
═══════════════════════════════════════════════════════
1. Si NO es factura DIAN válida → responde EXACTAMENTE: NO_ES_FACTURA
2. Si SÍ es factura DIAN → extrae TODOS los campos posibles aunque sean parciales.
3. Números: decimal con punto (ej: 1234567.50). Sin puntos de miles. Si no aparece → 0.
4. Texto: valor exacto tal como aparece. Si no aparece → "N/A".
5. numero_factura: 'No. Factura', 'Número', 'FACT-', 'FE-', 'FV', 'FES', o UUID/CUFE en XML.
6. proveedor: quien EMITE (vende). cliente: quien RECIBE (compra/paga).
7. numero_id: NIT sin dígito verificador (ej: '900123456' no '900123456-1').
8. iva: suma de TaxAmount con TaxCode=01 o nombre 'IVA'.
9. rete_iva: retención sobre IVA (normalmente 15% del IVA).
10. rete_ica: Impuesto de Industria y Comercio retenido.
11. retencion_fuente: busca 'RteFte', 'RetFte', 'Retención fuente'.
12. valor_total: valor final a pagar tras impuestos y retenciones.
13. estado: PAGADA si aparece 'pagado/cancelado/recibido', VENCIDA si venció sin pagar → sino PENDIENTE.
14. clasificacion: Servicios/Productos/Mixto según ítems.
15. En XML DIAN UBL 2.1: LineExtensionAmount=subtotal, PayableAmount=valor_total.

Respóndeme ÚNICAMENTE con el JSON o con NO_ES_FACTURA. Sin texto antes ni después, sin bloques ```:
{JSON_SCHEMA}

CONTENIDO:
{full_text}
"""

        content: list = []
        if image_parts:
            for media_type, b64, fname in image_parts[:5]:
                content.append({"type": "text", "text": f"[Imagen adjunta: {fname}]"})
                content.append({"type": "image", "source": {
                    "type": "base64",
                    "media_type": media_type,
                    "data": b64
                }})
        content.append({"type": "text", "text": prompt})

        resp = client.messages.create(
            model=CLAUDE_MODEL,
            max_tokens=4096,
            messages=[{"role": "user", "content": content}]
        )
        text = resp.content[0].text.strip()

        if "NO_ES_FACTURA" in text:
            return None

        def _extract_json(raw: str) -> Optional[dict]:
            try:
                return json.loads(raw)
            except Exception:
                pass
            block = re.search(r'```(?:json)?\s*([\s\S]*?)```', raw)
            if block:
                try:
                    return json.loads(block.group(1).strip())
                except Exception:
                    pass
            start = raw.find('{')
            if start == -1:
                return None
            depth = 0
            for i, ch in enumerate(raw[start:], start):
                if ch == '{':
                    depth += 1
                elif ch == '}':
                    depth -= 1
                    if depth == 0:
                        try:
                            return json.loads(raw[start:i+1])
                        except Exception:
                            return None
            return None

        data = _extract_json(text)
        if data:
            return data
        print(f"⚠️ Claude no devolvió JSON válido. Respuesta: {text[:200]}")
        return None

      except Exception as e:
        err_str = str(e)
        if "429" in err_str or "rate_limit" in err_str:
            # Backoff exponencial: 60 s, 120 s, 180 s
            wait = 60 * (attempt + 1)
            print(f"⏳ Rate limit Anthropic (intento {attempt+1}/3). Esperando {wait}s...")
            _time.sleep(wait)
            continue   # reintentar
        print(f"⚠️ Error IA: {e}")
        return None

    print(f"⚠️ Error IA: Se agotaron los reintentos por rate limit en '{subject[:50]}'")
    return None


def _is_duplicate_invoice(numero_factura: str, mes_nombre: str = None,
                           proveedor: str = None, fecha: str = None) -> tuple:
    """Busca la factura en TODAS las hojas mensuales."""
    num_clean = str(numero_factura).strip() if numero_factura else ""
    use_num   = bool(num_clean and num_clean not in ("", "N/A"))
    prov_clean  = str(proveedor or "").strip().lower()
    fecha_clean = str(fecha or "").strip()
    use_composite = (not use_num) and bool(prov_clean and prov_clean != "n/a")
    if not use_num and not use_composite:
        return False, -1, None

    hojas_revisadas = set()
    try:
        all_ws = get_all_monthly_worksheets()
        for ws in all_ws:
            hojas_revisadas.add(ws.title)
            try:
                if use_num:
                    col_values = _api_call_with_retry(ws.col_values, 3)
                    for i, val in enumerate(col_values[1:], start=2):
                        if str(val).strip() == num_clean:
                            return True, i, ws
                elif use_composite:
                    col_fecha = _api_call_with_retry(ws.col_values, 2)
                    col_prov  = _api_call_with_retry(ws.col_values, 4)
                    for i in range(1, max(len(col_fecha), len(col_prov))):
                        v_prov  = str(col_prov[i]).strip().lower()  if i < len(col_prov)  else ""
                        v_fecha = str(col_fecha[i]).strip()         if i < len(col_fecha) else ""
                        if v_prov == prov_clean and (not fecha_clean or v_fecha == fecha_clean):
                            return True, i + 1, ws
            except Exception:
                continue
        if mes_nombre and mes_nombre not in hojas_revisadas:
            try:
                ws = get_sheet(mes_nombre)
                if use_num:
                    col_values = _api_call_with_retry(ws.col_values, 3)
                    for i, val in enumerate(col_values[1:], start=2):
                        if str(val).strip() == num_clean:
                            return True, i, ws
                elif use_composite:
                    col_fecha = _api_call_with_retry(ws.col_values, 2)
                    col_prov  = _api_call_with_retry(ws.col_values, 4)
                    for i in range(1, max(len(col_fecha), len(col_prov))):
                        v_prov  = str(col_prov[i]).strip().lower()  if i < len(col_prov)  else ""
                        v_fecha = str(col_fecha[i]).strip()         if i < len(col_fecha) else ""
                        if v_prov == prov_clean and (not fecha_clean or v_fecha == fecha_clean):
                            return True, i + 1, ws
            except Exception:
                pass
        return False, -1, None
    except Exception:
        return False, -1, None


def _safe_float(val) -> float:
    """Convierte a float soportando formato colombiano (coma decimal, punto miles)."""
    try:
        s = str(val).strip()
        if not s or s in ("", "N/A", "0"):
            return 0.0
        if "," in s and "." in s:
            s = s.replace(".", "").replace(",", ".")
        elif "," in s:
            parts = s.split(",")
            if len(parts) == 2 and len(parts[1]) <= 2:
                s = s.replace(",", ".")
            else:
                s = s.replace(",", "")
        return float(s)
    except Exception:
        return 0.0


def _build_invoice_row(invoice_data: dict, mes_nombre: str) -> list:
    """Construye la lista de valores en el mismo orden que COLUMNS."""
    valor_total  = _safe_float(invoice_data.get("valor_total")  or 0)
    valor_pagado = _safe_float(invoice_data.get("valor_pagado") or 0)
    estado       = str(invoice_data.get("estado") or "PENDIENTE").strip().upper()

    if estado == "PAGADA" and valor_pagado == 0 and valor_total > 0:
        valor_pagado = valor_total

    valor_por_pagar = max(0.0, valor_total - valor_pagado)

    return [
        mes_nombre,
        str(invoice_data.get("fecha_factura") or "N/A"),
        str(invoice_data.get("numero_factura") or "N/A"),
        str(invoice_data.get("proveedor")      or "N/A"),
        str(invoice_data.get("id_tipo")        or "N/A"),
        str(invoice_data.get("numero_id")      or "N/A"),
        _safe_float(invoice_data.get("subtotal")          or 0),
        _safe_float(invoice_data.get("descuento")         or 0),
        _safe_float(invoice_data.get("iva")               or 0),
        _safe_float(invoice_data.get("rete_iva")          or 0),
        _safe_float(invoice_data.get("rete_ica")          or 0),
        _safe_float(invoice_data.get("impto_consumo")     or 0),
        _safe_float(invoice_data.get("propina")           or 0),
        _safe_float(invoice_data.get("otros_impuestos")   or 0),
        _safe_float(invoice_data.get("retencion_fuente")  or 0),
        _safe_float(invoice_data.get("administracion")    or 0),
        _safe_float(invoice_data.get("utilidad")          or 0),
        _safe_float(invoice_data.get("imprevistos")       or 0),
        valor_total,
        str(invoice_data.get("clasificacion") or "N/A"),
        str(estado),
        valor_pagado,
        valor_por_pagar,
        str(invoice_data.get("fecha_pago")    or "N/A"),
        str(invoice_data.get("cliente")       or "N/A"),
        str(invoice_data.get("cotizacion")    or "N/A"),
        str(invoice_data.get("observaciones") or ""),
    ]


# ── GUARDAR EN GOOGLE SHEETS ──────────────────────────────────────────────────
def save_to_sheets(invoice_data: dict, email_from: str, forced_mes: str = None):
    """INSERT si la factura no existe, UPDATE si ya existe con datos incompletos."""
    try:
        numero    = str(invoice_data.get("numero_factura") or "N/A")
        proveedor = str(invoice_data.get("proveedor")      or "N/A")

        meses_list = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
                      "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
        if forced_mes:
            mes_nombre = forced_mes.strip().capitalize()
        else:
            fecha_factura_str = str(invoice_data.get("fecha_factura") or "N/A")
            try:
                if "/" in fecha_factura_str:
                    mes_num    = int(fecha_factura_str.split("/")[1])
                    mes_nombre = meses_list[mes_num - 1] if 1 <= mes_num <= 12 else meses_list[datetime.now().month - 1]
                else:
                    mes_nombre = meses_list[datetime.now().month - 1]
            except Exception:
                mes_nombre = meses_list[datetime.now().month - 1]

        fecha_factura_str = str(invoice_data.get("fecha_factura") or "N/A")

        found, dup_row, dup_ws = _is_duplicate_invoice(
            numero, mes_nombre, proveedor, fecha_factura_str
        )

        if found and dup_ws is not None:
            try:
                existing_row_vals = _api_call_with_retry(dup_ws.row_values, dup_row)
                existing_total  = _safe_float(existing_row_vals[18]) if len(existing_row_vals) > 18 and existing_row_vals[18] not in ("", "0", "N/A") else 0.0
                existing_estado = str(existing_row_vals[20]).strip().upper() if len(existing_row_vals) > 20 and existing_row_vals[20] else "PENDIENTE"
                new_total       = _safe_float(invoice_data.get("valor_total") or 0)
                new_estado      = str(invoice_data.get("estado") or "PENDIENTE").strip().upper()

                estado_mejoro = (existing_estado in ("PENDIENTE", "VENCIDA") and new_estado == "PAGADA")
                datos_mejoran = (new_total > 0 and existing_total == 0)
                total_cambio  = (new_total > 0 and existing_total > 0 and abs(new_total - existing_total) > 1)

                if datos_mejoran or estado_mejoro or total_cambio:
                    new_row = _build_invoice_row(invoice_data, mes_nombre)
                    for col_idx, val in enumerate(new_row, start=1):
                        _api_call_with_retry(dup_ws.update_cell, dup_row, col_idx, val)
                    razon = "datos completos" if datos_mejoran else ("estado→PAGADA" if estado_mejoro else "importe corregido")
                    print(f"🔄 Factura ACTUALIZADA ({razon}): {numero} — {proveedor}")
                    return True
                else:
                    print(f"⏭️  Factura sin cambios, omitida: {numero} — {proveedor} ({existing_estado})")
                    return False
            except Exception as e_upd:
                print(f"⚠️  No se pudo procesar duplicado {numero}: {e_upd}")
                return False

        row = _build_invoice_row(invoice_data, mes_nombre)
        ws  = get_sheet(mes_nombre)
        _api_call_with_retry(ws.append_row, row, value_input_option="USER_ENTERED")
        print(f"✅ Factura guardada en hoja '{mes_nombre}': {numero} — {proveedor}")
        return True

    except Exception as e:
        print(f"❌ Error guardando en Sheets: {e}")
        import traceback
        print(traceback.format_exc())
        return False


def update_invoice_status(numero_factura: str, nuevo_estado: str,
                          valor_pagado: float = None, fecha_pago: str = None,
                          observaciones: str = None) -> dict:
    """Actualiza Estado, Valor Pagado, Valor Por Pagar, Fecha Pago y Observaciones
    de una factura existente en Google Sheets.

    Busca la factura en todas las hojas mensuales por número de factura.
    Retorna {"ok": True, "mes": "Febrero", "fila": 5} o {"ok": False, "error": "..."}.

    Índices de columna (1-based, según COLUMNS):
      19 → Valor Total        21 → Estado
      22 → Valor Pagado       23 → Valor Por Pagar
      24 → Fecha Pago         27 → Observaciones
    """
    num_clean = str(numero_factura).strip()
    if not num_clean or num_clean == "N/A":
        return {"ok": False, "error": "Número de factura inválido"}

    nuevo_estado = nuevo_estado.strip().upper()
    if nuevo_estado not in ("PENDIENTE", "PAGADA", "VENCIDA"):
        return {"ok": False, "error": f"Estado inválido: {nuevo_estado}"}

    try:
        found, row_num, ws = _is_duplicate_invoice(num_clean)
        if not found or ws is None:
            return {"ok": False, "error": f"Factura '{num_clean}' no encontrada"}

        # Leer fila actual para calcular Valor Por Pagar
        row_vals = _api_call_with_retry(ws.row_values, row_num)
        valor_total = _safe_float(row_vals[18]) if len(row_vals) > 18 else 0.0  # col 19

        # Calcular valor_pagado si no se proporcionó
        if valor_pagado is None:
            if nuevo_estado == "PAGADA":
                valor_pagado = valor_total
            else:
                valor_pagado = _safe_float(row_vals[21]) if len(row_vals) > 21 else 0.0  # col 22

        valor_por_pagar = max(0.0, valor_total - valor_pagado)

        # Si marca PAGADA y no hay fecha, poner la de hoy
        if fecha_pago is None and nuevo_estado == "PAGADA":
            fecha_pago = datetime.now().strftime("%d/%m/%Y")
        elif fecha_pago is None:
            fecha_pago = row_vals[23] if len(row_vals) > 23 else "N/A"  # col 24

        # Observaciones: append si hay texto nuevo, conservar si no
        obs_actual = row_vals[26] if len(row_vals) > 26 else ""  # col 27
        if observaciones and observaciones.strip():
            obs_nueva = f"{obs_actual} | {observaciones.strip()}".lstrip(" |")
        else:
            obs_nueva = obs_actual

        # Actualizar celdas
        updates = {
            21: nuevo_estado,
            22: valor_pagado,
            23: valor_por_pagar,
            24: fecha_pago,
            27: obs_nueva,
        }
        for col_idx, val in updates.items():
            _api_call_with_retry(ws.update_cell, row_num, col_idx, val)

        print(f"✅ Estado actualizado: {num_clean} → {nuevo_estado} (fila {row_num}, hoja '{ws.title}')")
        return {"ok": True, "mes": ws.title, "fila": row_num,
                "estado": nuevo_estado, "valor_pagado": valor_pagado,
                "valor_por_pagar": valor_por_pagar, "fecha_pago": fecha_pago}

    except Exception as e:
        import traceback
        print(f"❌ Error actualizando factura {num_clean}: {e}")
        print(traceback.format_exc())
        return {"ok": False, "error": str(e)}


def export_to_excel() -> str:
    """Descarga todos los registros de todas las hojas mensuales y genera un Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    try:
        spreadsheet = _get_spreadsheet()
        worksheets = _api_call_with_retry(spreadsheet.worksheets)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        for ws_gsheet in worksheets:
            if ws_gsheet.title in MESES:
                rows = ws_gsheet.get_all_values()
                if not rows:
                    continue

                sheet = wb.create_sheet(title=ws_gsheet.title)
                for i, row in enumerate(rows, 1):
                    sheet.append(row)
                    if i == 1:
                        for cell in sheet[1]:
                            cell.font = Font(bold=True, color="FFFFFF")
                            cell.fill = PatternFill("solid", fgColor="1F4E79")
                            cell.alignment = Alignment(horizontal="center")

        if not wb.worksheets:
            sheet = wb.create_sheet(title="Sin Datos")
            sheet.append(["No hay datos disponibles"])

        wb.save(EXCEL_PATH)
        print(f"✅ Excel exportado con {len(wb.worksheets)} hojas: {EXCEL_PATH}")
        return EXCEL_PATH
    except Exception as e:
        print(f"❌ Error exportando Excel: {e}")
        raise


# ── PROCESO PRINCIPAL ──────────────────────────────────────────────────────────
def process_emails():
    """Función principal del agente. Procesa los correos del mes actual."""
    mes_actual = ["enero","febrero","marzo","abril","mayo","junio",
                  "julio","agosto","septiembre","octubre","noviembre","diciembre"][datetime.now().month - 1]
    year_actual = datetime.now().year
    print(f"📅 process_emails(): delegando a mes actual → {mes_actual.capitalize()} {year_actual}")
    process_emails_for_month(mes_actual, year_actual)


# ── ENTRY POINT STANDALONE ─────────────────────────────────────────────────────
if __name__ == "__main__":
    process_emails()
